from openpyxl import load_workbook, Workbook

wb = load_workbook("sales_data.xlsx")

sheet = wb.active

sheet.cell(row=1, column=4).value = "Total"

for row in range(2, sheet.max_row + 1):
    quantity = sheet.cell(row=row, column=2).value
    price = sheet.cell(row=row, column=3).value
    sheet.cell(row=row, column=4).value = quantity * price

new_wb = Workbook()
new_sheet = new_wb.active
new_sheet.title = "Summary"

for row in sheet.iter_rows(values_only=True):
    new_sheet.append(row)

new_wb.save("sales_summary.xlsx")

print("Excel processed without sheet name!")
